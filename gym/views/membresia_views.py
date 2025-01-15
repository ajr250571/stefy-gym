from django.db.models.base import Model as Model
from django.shortcuts import redirect
from django.urls import reverse_lazy
from gym.models import Membresia, Asistencia, Plan
from gym.filters import MembresiaFilter
from gym.forms import MembresiaForm
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, DetailView, DetailView
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime
from django_filters.views import FilterView
from dateutil.relativedelta import relativedelta
from django.utils import timezone


class membresiaListView(PermissionRequiredMixin, FilterView):
    model = Membresia
    template_name = 'membresia/membresia_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_membresia'
    filterset_class = MembresiaFilter

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Membresias'
        kwargs['create_url'] = reverse_lazy('membresia_create')
        kwargs['crumb_url'] = reverse_lazy('membresia_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        # 'ACTIVA'
        return Membresia.objects.all()

    def export_excel(self, request):
        # Crear un nuevo libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Membresias"  # type: ignore

        # Definir encabezados
        headers = ['Socio', 'Plan', 'Fecha Inicio', 'Fecha Fin', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)  # type: ignore
            cell.value = header
            cell.font = Font(bold=True)

        # Obtener los datos filtrados
        queryset = self.filterset.qs if hasattr(
            self, 'filterset') else self.get_queryset()

        # Llenar datos
        for row, membresia in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=str(  # type: ignore
                membresia.socio))  # type: ignore
            ws.cell(row=row, column=2, value=str(  # type: ignore
                membresia.plan))  # type: ignore
            ws.cell(row=row, column=3,  # type: ignore
                    value=membresia.fecha_inicio.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=4,  # type: ignore
                    value=membresia.fecha_fin.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=membresia.estado)  # type: ignore

        # Ajustar anchos de columna
        for column in ws.columns:  # type: ignore
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(  # type: ignore
                column[0].column)  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Membresias_{}.xlsx'.format(
            datetime.now().strftime('%Y%m%d_%H%M%S')
        )

        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_excel(request)
        return super().get(request, *args, **kwargs)

# Listar membresias vencidas


class membresiaActivaListView(PermissionRequiredMixin, ListView):
    model = Membresia
    template_name = 'membresia/membresia_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Membresias Activas'
        kwargs['create_url'] = reverse_lazy('membresia_create')
        kwargs['crumb_url'] = reverse_lazy('membresia_activa_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        # 'ACTIVA'
        return Membresia.objects.filter(estado='ACTIVA')


class membresiaVencidaListView(PermissionRequiredMixin, ListView):
    model = Membresia
    template_name = 'membresia/membresia_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Membresias Vencidas'
        kwargs['create_url'] = reverse_lazy('membresia_create')
        kwargs['crumb_url'] = reverse_lazy('membresia_vencida_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        # 'VENCIDA'
        return Membresia.objects.filter(estado='VENCIDA')


class membresiaCanceladaListView(PermissionRequiredMixin, ListView):
    model = Membresia
    template_name = 'membresia/membresia_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Membresias Canceladas'
        kwargs['create_url'] = reverse_lazy('membresia_create')
        kwargs['crumb_url'] = reverse_lazy('membresia_cancelada_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        # 'CANCELADA'
        return Membresia.objects.filter(estado='CANCELADA')


class membresiaDetailDniView(DetailView):
    model = Membresia
    template_name = 'membresia/membresia_detail.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Detalle de Membresia'
        kwargs['crumb_url'] = reverse_lazy('home')
        kwargs['crumb_name'] = ''
        return super().get_context_data(**kwargs)

    def get_object(self, queryset=None):
        dni = self.kwargs['dni']
        try:
            membresia = Membresia.objects.get(socio__dni=dni)
            Asistencia.registrar_asistencia(membresia.socio.dni)
            return membresia
        except Exception as e:
            print(f"Error al obtener membresía: {str(e)}")
            return redirect('home')


class membresiaCreateView(PermissionRequiredMixin, CreateView):
    model = Membresia
    form_class = MembresiaForm
    template_name = 'membresia/membresia_create.html'
    success_url = reverse_lazy('membresia_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.add_membresia'

    def get_initial(self):
        self.initial['fecha_fin'] = timezone.now().date() + \
            relativedelta(months=1)
        return super().get_initial()

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Crear Membresia'
        kwargs['crumb_url'] = reverse_lazy('membresia_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)


class membresiaSocioCreateView(PermissionRequiredMixin, CreateView):
    model = Membresia
    form_class = MembresiaForm
    template_name = 'membresia/membresia_create.html'
    success_url = reverse_lazy('membresia_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.add_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Crear Membresia'
        kwargs['crumb_url'] = reverse_lazy('membresia_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)

    def get_initial(self):

        initial = super().get_initial()

        initial['socio'] = self.kwargs['pk']
        initial['fecha_fin'] = timezone.now().date() + relativedelta(months=1)
        # Asigna plan por defecto al socio
        plan = Plan.objects.get(duracion=1)
        initial['plan'] = plan

        return initial


class membresiaUpdateView(PermissionRequiredMixin, UpdateView):
    model = Membresia
    form_class = MembresiaForm
    template_name = 'membresia/membresia_create.html'
    success_url = reverse_lazy('membresia_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.change_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Editar Membresia'
        kwargs['crumb_url'] = reverse_lazy('membresia_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)


class membresiaDeleteView(PermissionRequiredMixin, DeleteView):
    model = Membresia
    template_name = 'confirm_delete.html'
    success_url = reverse_lazy('membresia_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.delete_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['cancel_url'] = reverse_lazy('membresia_list')
        return super().get_context_data(**kwargs)


class membresiaDetailView(PermissionRequiredMixin, DetailView):
    model = Membresia
    template_name = 'membresia/membresia_detail.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_membresia'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Detalle de Membresia'
        kwargs['crumb_url'] = reverse_lazy('membresia_list')
        kwargs['crumb_name'] = 'Membresias'
        return super().get_context_data(**kwargs)
