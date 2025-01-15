from django.db.models.base import Model as Model
from django.shortcuts import redirect
from django.urls import reverse_lazy, reverse
from gym.models import Membresia, Pago
from gym.filters import PagoFilter
from gym.forms import PagoForm
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView, DetailView
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime, timezone
from django_filters.views import FilterView
from dateutil.relativedelta import relativedelta
from datetime import date, timedelta, datetime
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse

# Pagos


class pagoListView(PermissionRequiredMixin, FilterView):
    model = Pago
    template_name = 'pago/pago_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_pago'
    filterset_class = PagoFilter

    def get_filterset(self, filterset_class):
        # Si no hay parámetros en la URL, establecemos el valor inicial
        if not self.request.GET:
            return filterset_class(
                data={'fecha_at': 'month'},
                queryset=self.get_queryset(),
                request=self.request
            )
        return super().get_filterset(filterset_class)

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Pagos'
        kwargs['create_url'] = reverse_lazy('pago_create')
        kwargs['crumb_url'] = reverse_lazy('pago_list')
        kwargs['crumb_name'] = 'Pagos'
        # kwargs['filter_form'] = FechaFilterForm()
        return super().get_context_data(**kwargs)

    def export_excel(self, request):
        # Crear un nuevo libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagos"

        # Definir encabezados
        headers = ['Membresía', 'Fecha', 'Monto', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Obtener los datos filtrados
        queryset = self.filterset.qs if hasattr(
            self, 'filterset') else self.get_queryset()

        # Llenar datos
        for row, pago in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=str(pago.membresia))
            ws.cell(row=row, column=2, value=pago.fecha_pago.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=float(pago.monto))
            ws.cell(row=row, column=4, value=pago.estado)

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Pagos_{}.xlsx'.format(
            datetime.now().strftime('%Y%m%d_%H%M%S')
        )

        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_excel(request)
        return super().get(request, *args, **kwargs)


class pagoCreateView(PermissionRequiredMixin, CreateView):
    model = Pago
    form_class = PagoForm
    template_name = 'pago/pago_create.html'
    success_url = reverse_lazy('pago_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.add_pago'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Crear Pago'
        kwargs['crumb_url'] = reverse_lazy('pago_list')
        kwargs['crumb_name'] = 'Pagos'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        try:
            # Asegurarse de que el monto se establezca correctamente
            membresia = form.cleaned_data.get('membresia')
            if membresia:
                form.instance.monto = membresia.plan.precio

            response = super().form_valid(form)
            messages.success(self.request, 'Pago registrado exitosamente.')
            return response
        except Exception as e:
            messages.error(self.request, f'Error al guardar el pago: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Imprimir errores en la consola para debug
        print("Errores del formulario:", form.errors)
        print("Datos enviados:", form.data)
        messages.error(
            self.request, f'Errores en el formulario: {form.errors}')
        return super().form_invalid(form)

    def form_valid(self, form):
        try:
            # Asegurarse de que todos los campos necesarios estén presentes
            membresia = form.cleaned_data.get('membresia')
            if not membresia:
                form.add_error('membresia', 'Este campo es requerido')
                return self.form_invalid(form)

            # Establecer los campos calculados
            form.instance.monto = membresia.plan.precio
            form.instance.fecha_vencimiento = membresia.fecha_fin

            response = super().form_valid(form)
            messages.success(self.request, 'Pago registrado exitosamente.')
            return response
        except Exception as e:
            print("Error al guardar:", str(e))
            messages.error(self.request, f'Error al guardar el pago: {str(e)}')
            return self.form_invalid(form)


class pagoMembresiaCreateView(PermissionRequiredMixin, CreateView):
    model = Pago
    form_class = PagoForm
    template_name = 'pago/pago_create.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.add_pago'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_success_url(self):
        # Obtener parámetros de la URL
        param = self.kwargs.get('pk')
        if param:
            # Si hay parámetro, redirigir a una URL específica
            return reverse('membresia_list')
        else:
            # Si no hay parámetro, redirigir a la URL por defecto
            return reverse('pago_list')

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Crear Pago'
        kwargs['crumb_url'] = reverse_lazy('pago_list')
        kwargs['crumb_name'] = 'Pagos'
        return super().get_context_data(**kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial['membresia'] = self.kwargs['pk']
        membresia = Membresia.objects.get(pk=self.kwargs['pk'])
        if membresia.plan == None:
            initial['monto'] = 0
        else:
            initial['monto'] = membresia.plan.precio
            initial['fecha_vencimiento'] = membresia.fecha_fin

        initial['fecha_pago'] = date.today()
        initial['estado'] = 'PAGADO'
        return initial


class pagoUpdateView(PermissionRequiredMixin, UpdateView):
    model = Pago
    form_class = PagoForm
    template_name = 'pago/pago_create.html'
    success_url = reverse_lazy('pago_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.change_pago'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Editar Pago'
        kwargs['crumb_url'] = reverse_lazy('pago_list')
        kwargs['crumb_name'] = 'Pagos'
        return super().get_context_data(**kwargs)


class pagoDeleteView(PermissionRequiredMixin, DeleteView):
    model = Pago
    template_name = 'confirm_delete.html'
    success_url = reverse_lazy('pago_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.delete_pago'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['cancel_url'] = reverse_lazy('pago_list')
        return super().get_context_data(**kwargs)


class errorPermisosView(TemplateView):
    template_name = 'error_permisos.html'


class MontosMensualesView(TemplateView):
    template_name = 'pago/montos_mensuales.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado Montos Mensuales'
        montos_mensuales = Pago.objects.annotate(
            mes=TruncMonth('fecha_pago')
        ).values('mes').annotate(
            total=Sum('monto')
        ).order_by('-mes')

        context['montos_mensuales'] = montos_mensuales
        return context


def get_membresia_monto(request):
    """Vista para obtener el monto del plan de una membresía vía AJAX"""
    membresia_id = request.GET.get('membresia_id')
    try:
        membresia = Membresia.objects.get(id=membresia_id)
        return JsonResponse({
            'monto': float(membresia.plan.precio),
            'fecha_vencimiento': membresia.fecha_fin,
            'success': True
        })
    except Membresia.DoesNotExist:
        return JsonResponse({
            'success': False,
            'fecha_vencimiento': membresia.fecha_fin,
            'error': 'Membresía no encontrada'
        })


class pagoDetailView(PermissionRequiredMixin, DetailView):
    model = Pago
    template_name = 'pago/pago_detail.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_pago'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Detalle de Pago'
        kwargs['crumb_url'] = reverse_lazy('pago_list')
        kwargs['crumb_name'] = 'Pagos'
        return super().get_context_data(**kwargs)
