from django.db.models.base import Model as Model
from django.shortcuts import redirect
from django.urls import reverse_lazy, reverse
from gym.models import Socio
from gym.forms import SocioForm
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, DetailView
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
from gym.filters import SocioFilter

# Socios


class socioListView(PermissionRequiredMixin, ListView):
    model = Socio
    template_name = 'socio/socio_list.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_socio'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Listado de Socios'
        kwargs['create_url'] = reverse_lazy('socio_create')
        kwargs['crumb_url'] = reverse_lazy('socio_list')
        kwargs['crumb_name'] = 'Socios'
        return super().get_context_data(**kwargs)

    def export_excel(self, request):
        # Crear un nuevo libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Socios"  # type: ignore

        # Definir encabezados
        headers = ['Socio', 'DNI', 'Telefono', 'Email', 'Direccion', 'Activo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)  # type: ignore
            cell.value = header
            cell.font = Font(bold=True)

        # Obtener los datos filtrados
        queryset = self.filterset.qs if hasattr(  # type: ignore
            self, 'filterset') else self.get_queryset()

        # Llenar datos
        for row, membresia in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=str(  # type: ignore
                membresia.nombre_completo()))  # type: ignore
            ws.cell(row=row, column=2, value=str(  # type: ignore
                membresia.dni))  # type: ignore
            ws.cell(row=row, column=3, value=str(  # type: ignore
                membresia.telefono))  # type: ignore
            ws.cell(row=row, column=4, value=str(  # type: ignore
                membresia.email))  # type: ignore
            ws.cell(row=row, column=5, value=str(  # type: ignore
                membresia.direccion))  # type: ignore
            ws.cell(row=row, column=6, value=str(  # type: ignore
                membresia.activo))  # type: ignore

        # Ajustar anchos de columna
        for column in ws.columns:  # type: ignore
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(  # type: ignore
                column[0].column)  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # type: ignore
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Socios_{}.xlsx'.format(
            datetime.now().strftime('%Y%m%d_%H%M%S')
        )

        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_excel(request)
        return super().get(request, *args, **kwargs)


class socioCreateView(PermissionRequiredMixin, CreateView):
    model = Socio
    form_class = SocioForm
    template_name = 'socio/socio_create.html'
    # success_url = reverse_lazy('socio_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.add_socio'

    def get_success_url(self) -> str:
        # Redirigir a crete Membresia si el socio fue creado exitosamente
        return reverse('membresia_socio_create',
                       kwargs={'pk': self.object.pk})  # type: ignore

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Crear Socio'
        kwargs['crumb_url'] = reverse_lazy('socio_list')
        kwargs['crumb_name'] = 'Socios'
        return super().get_context_data(**kwargs)


class socioUpdateView(PermissionRequiredMixin, UpdateView):
    model = Socio
    form_class = SocioForm
    template_name = 'socio/socio_create.html'
    success_url = reverse_lazy('socio_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.change_socio'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Editar Socio'
        kwargs['crumb_url'] = reverse_lazy('socio_list')
        kwargs['crumb_name'] = 'Socios'
        return super().get_context_data(**kwargs)


class socioDeleteView(PermissionRequiredMixin, DeleteView):
    model = Socio
    template_name = 'confirm_delete.html'
    success_url = reverse_lazy('socio_list')
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.delete_socio'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['cancel_url'] = reverse_lazy('socio_list')
        return super().get_context_data(**kwargs)


class socioDetailView(PermissionRequiredMixin, DetailView):
    model = Socio
    template_name = 'socio/socio_detail.html'
    login_url = '/login/'
    permisos_url = '/error_permisos/'
    permission_required = 'gym.view_socio'

    def handle_no_permission(self):
        messages.error(
            self.request, 'No tienes permisos para realizar esta acción.')
        return redirect(self.permisos_url)

    def get_context_data(self, **kwargs):
        kwargs['title'] = 'Detalle Socio'
        kwargs['crumb_url'] = reverse_lazy('socio_list')
        kwargs['crumb_name'] = 'Socios'
        kwargs['cancel_url'] = reverse_lazy('socio_list')
        return super().get_context_data(**kwargs)
